#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2023/12/11 21:24
# @Author  : 刘双喜
# @File    : 任意倍组合数.py
# @Description : 添加描述
import os.path
from openpyxl import load_workbook
from openpyxl import Workbook
import pandas as pd

if os.path.exists('input1.xlsx'):
    pass
else:
    data = Workbook()  # 创建
    # 创建工作表
    data.create_sheet('Sheet')
    data['Sheet']['A1'].value = '单价'
    data.save("input1.xlsx")  # 保存
input('在input1的第一列复制单价,完成后保存退出')
num = pd.read_excel('input1.xlsx', 'Sheet')
num = num.values.tolist()
num = [i[0] for i in num]
num = list(map(float, num))
limit = int(input('输入最大限制值'))
num = sorted(list(set(num)))
res = []
cal = []
# 选取第i个数
for i in num:
    # 第i个数的所有结果
    res_i = [round(i * idx, 5) for idx in range(1, 1 + int(limit / i))]
    cal_i = [f'{i}*{idx}' for idx in range(1, 1 + int(limit / i))]
    if res:
        temp = []
        temp_cal = []
        # 叠加已有的结果
        for j in res_i:
            temp += [round(j + _, 5) for _ in res if j + _ <= limit]
            temp_cal += [f'{j}+{_}' for _ in res if j + _ <= limit]
        res += temp
        cal += temp_cal
    else:
        res += res_i
        cal += cal_i

res = dict(zip(res, cal))
res = {i: res.get(i) for i in sorted(res)}
date = load_workbook("input1.xlsx")  # 加载
sheet = date['Sheet1']
for idx, key in enumerate(res.keys(), 1):
    sheet[f'A{idx}'].value = key
    sheet[f'B{idx}'].value = res.get(key)
date.save('input1.xlsx')
