#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2023/11/11 23:09
# @Author  : 刘双喜
# @File    : HJ0.py
# @Description : 添加描述
from openpyxl import load_workbook
import os

with load_workbook('1.xlsx') as excel:
    sheet_name = excel.sheetnames[0]
company_name = sheet_name.split('-')[0]
os. rename(old, new)
